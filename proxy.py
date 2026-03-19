import os
from flask import Flask, send_from_directory, request, jsonify
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from openpyxl import load_workbook
import io

app = Flask(__name__)
# Database will be created in the same folder
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///prohorizon_tasks.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# Task Database Model
class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(200), nullable=False)
    assignee = db.Column(db.String(100))
    assigned_by = db.Column(db.String(100))
    priority = db.Column(db.String(20)) 
    status = db.Column(db.String(50), default='Pending') # Increased length for custom statuses
    deadline = db.Column(db.String(20))
    remark = db.Column(db.Text, default='')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# Initialize Database
with app.app_context():
    db.create_all()

@app.route('/')
def index():
    # Serves your updated index.html from the same directory
    return send_from_directory('.', 'index.html')

# Fetch and Create Tasks
@app.route('/api/tasks', methods=['GET', 'POST'])
def manage_tasks():
    if request.method == 'POST':
        data = request.json
        new_task = Task(
            subject=data.get('subject'),
            assignee=data.get('assignee'),
            assigned_by=data.get('assigned_by'),
            priority=data.get('priority'),
            status=data.get('status', 'Pending'), # Now captures custom status on creation
            remark=data.get('remark', ''),        # Now captures remark on creation
            deadline=data.get('deadline')
        )
        db.session.add(new_task)
        db.session.commit()
        return jsonify({"message": "Success"})
    
    tasks = Task.query.order_by(Task.created_at.desc()).all()
    output = []
    for t in tasks:
        output.append({
            "id": t.id, 
            "subject": t.subject, 
            "assignee": t.assignee,
            "assigned_by": t.assigned_by, 
            "priority": t.priority,
            "status": t.status, 
            "deadline": t.deadline,
            "remark": t.remark, 
            "created_at": t.created_at.strftime("%Y-%m-%d %H:%M") # Format for JS filtering
        })
    return jsonify(output)

# Bulk Import from Excel
@app.route('/api/tasks/import', methods=['POST'])
def import_tasks():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    try:
        wb = load_workbook(file)
        sheet = wb.active
        count = 0

        # Skips header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue # Skip empty rows
            
            new_task = Task(
                subject=str(row[0]),
                assignee=str(row[1]),
                assigned_by=str(row[2]),
                priority=str(row[3]),
                status=str(row[4] if len(row) > 4 and row[4] else 'Pending'), # Support custom status in Excel
                remark=str(row[5] if len(row) > 5 and row[5] else ''),         # Support remarks in Excel
                deadline=str(row[6] if len(row) > 6 else '')
            )
            db.session.add(new_task)
            count += 1

        db.session.commit()
        return jsonify({"message": f"Successfully imported {count} tasks"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Update Any Field (Subject, Priority, Status, Remark)
@app.route('/api/tasks/update', methods=['POST'])
def update_task():
    data = request.json
    task = Task.query.get(data['id'])
    if not task: return jsonify({"error": "Not found"}), 404
    
    # Dynamically update any field passed from the frontend
    if 'subject' in data: task.subject = data['subject']
    if 'remark' in data: task.remark = data['remark']
    if 'status' in data: task.status = data['status']
    if 'priority' in data: task.priority = data['priority']
    
    db.session.commit()
    return jsonify({"success": True})

if __name__ == '__main__':
    # Using environment port for deployment, defaults to 8000 for local/office server
    port = int(os.environ.get("PORT", 8000))
    app.run(host='0.0.0.0', port=port)
